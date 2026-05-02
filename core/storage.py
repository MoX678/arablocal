"""Storage layer: SQLite database, CSV output, and checkpoint management."""

from __future__ import annotations

import asyncio
import csv
import json
import logging
import os
import re
import sqlite3
import threading
import time
from datetime import datetime, timezone
from typing import Dict, List, Optional, Set

from core.config import BASE_FIELDS, SOCIAL_DOMAINS

log = logging.getLogger("arablocal")


class StorageManager:
    """Handles all persistence: SQLite staging DB, raw CSV append, and final CSV export.

    Each StorageManager owns one DB file, one raw CSV, and one output directory.
    Thread-safe via asyncio locks for concurrent producer/consumer access.
    """

    def __init__(self, db_path: str, raw_csv_path: str, categories_dir: str,
                 all_csv_path: str, start_time_ref: Optional[callable] = None):
        """
        Args:
            db_path: Path to SQLite staging database.
            raw_csv_path: Path to incremental raw CSV file.
            categories_dir: Directory for per-category exported CSVs.
            all_csv_path: Path for combined all-categories CSV.
            start_time_ref: Callable returning pipeline start time (for Runtime_Sec column).
        """
        self.db_path = db_path
        self.raw_csv_path = raw_csv_path
        self.categories_dir = categories_dir
        self.all_csv_path = all_csv_path
        self._start_time_ref = start_time_ref

        self.db_lock = asyncio.Lock()
        self.csv_lock = asyncio.Lock()
        # Serializes export passes (DB-based + raw fallback) across threads
        # so periodic exports can't race the final export at shutdown.
        self._export_lock = threading.Lock()

        # CSV batch buffer
        self._csv_buffer: List[dict] = []
        self._csv_fieldnames: Optional[List[str]] = None
        self._csv_flush_size = 25
        self._uses_stable_raw_header = os.path.basename(self.raw_csv_path).lower() in {"kw_raw.csv", "om_raw.csv"}

        self._init_db()

    # ─── Schema version ──────────────────────────────────────────────────

    CURRENT_SCHEMA_VERSION = 3

    # ─── Database Initialization ──────────────────────────────────────────

    def _init_db(self):
        with sqlite3.connect(self.db_path, timeout=15) as conn:
            # Enable WAL mode for concurrent reads during writes
            conn.execute("PRAGMA journal_mode=WAL")

            conn.execute("""
                CREATE TABLE IF NOT EXISTS businesses (
                    url TEXT PRIMARY KEY,
                    category TEXT NOT NULL,
                    data JSON NOT NULL,
                    scraped_at TEXT NOT NULL
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS categories (
                    slug TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    url TEXT NOT NULL,
                    discovered_at TEXT NOT NULL
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS checkpoints (
                    category_slug TEXT PRIMARY KEY,
                    last_page INTEGER NOT NULL DEFAULT 1,
                    urls_found INTEGER NOT NULL DEFAULT 0,
                    completed INTEGER NOT NULL DEFAULT 0,
                    updated_at TEXT NOT NULL
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS _meta (
                    key TEXT PRIMARY KEY,
                    value TEXT
                )
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_biz_category ON businesses(category)
            """)
            conn.commit()

            # Run schema migrations
            self._migrate(conn)

    def _get_schema_version(self, conn) -> int:
        """Get current schema version from _meta table."""
        try:
            row = conn.execute(
                "SELECT value FROM _meta WHERE key = 'schema_version'"
            ).fetchone()
            return int(row[0]) if row else 1
        except Exception:
            return 1

    def _set_schema_version(self, conn, version: int):
        conn.execute(
            "INSERT OR REPLACE INTO _meta (key, value) VALUES ('schema_version', ?)",
            (str(version),),
        )

    def _migrate(self, conn):
        """Run pending schema migrations sequentially."""
        version = self._get_schema_version(conn)
        if version >= self.CURRENT_SCHEMA_VERSION:
            return

        # Discover existing columns to make migrations idempotent
        existing_cols = {
            row[1] for row in conn.execute("PRAGMA table_info(businesses)")
        }

        # Migration v1 → v2: add delta + dedup columns
        if version < 2:
            log.info(f"[storage] Migrating schema v{version} → v2 (delta + dedup columns)")
            if "first_seen_at" not in existing_cols:
                conn.execute("ALTER TABLE businesses ADD COLUMN first_seen_at TEXT")
            if "last_seen_at" not in existing_cols:
                conn.execute("ALTER TABLE businesses ADD COLUMN last_seen_at TEXT")
            if "fingerprint" not in existing_cols:
                conn.execute("ALTER TABLE businesses ADD COLUMN fingerprint TEXT")
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_biz_fingerprint ON businesses(fingerprint)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_biz_last_seen ON businesses(last_seen_at)"
            )
            # Backfill existing rows
            conn.execute(
                "UPDATE businesses SET first_seen_at = scraped_at, "
                "last_seen_at = scraped_at WHERE first_seen_at IS NULL"
            )
            conn.commit()

        # Migration v2 → v3: clean bogus checkpoints (completed=1 but urls_found=0)
        if version < 3:
            cleaned = conn.execute(
                "DELETE FROM checkpoints WHERE completed = 1 AND urls_found = 0"
            ).rowcount
            conn.commit()
            if cleaned:
                log.info(f"[storage] Migration v3: cleaned {cleaned} bogus checkpoints")

        self._set_schema_version(conn, self.CURRENT_SCHEMA_VERSION)
        conn.commit()
        log.info(f"[storage] Schema at version {self.CURRENT_SCHEMA_VERSION}")

    # ─── Business CRUD ────────────────────────────────────────────────────

    async def insert_business(self, url: str, category: str, data: dict) -> str:
        """Insert or update a business record with delta tracking.

        Returns:
            "new" if this is the first time the URL was seen,
            "updated" if the URL already existed and was refreshed.
        """
        now = datetime.now(timezone.utc).isoformat()
        fingerprint = data.pop("_fingerprint", "") or ""
        async with self.db_lock:
            with sqlite3.connect(self.db_path, timeout=15) as conn:
                # Check if URL already exists
                existing = conn.execute(
                    "SELECT 1 FROM businesses WHERE url = ?", (url,)
                ).fetchone()

                conn.execute(
                    "INSERT INTO businesses "
                    "(url, category, data, scraped_at, first_seen_at, last_seen_at, fingerprint) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?) "
                    "ON CONFLICT(url) DO UPDATE SET "
                    "data = excluded.data, "
                    "last_seen_at = excluded.last_seen_at, "
                    "fingerprint = excluded.fingerprint",
                    (url, category, json.dumps(data, ensure_ascii=False),
                     now, now, now, fingerprint),
                )
                return "updated" if existing else "new"

    async def touch_last_seen(self, urls: List[str]):
        """Update last_seen_at for skipped URLs (already scraped)."""
        if not urls:
            return
        now = datetime.now(timezone.utc).isoformat()
        async with self.db_lock:
            with sqlite3.connect(self.db_path, timeout=15) as conn:
                conn.executemany(
                    "UPDATE businesses SET last_seen_at = ? WHERE url = ?",
                    [(now, u) for u in urls],
                )

    async def insert_category(self, slug: str, name: str, url: str):
        async with self.db_lock:
            with sqlite3.connect(self.db_path, timeout=15) as conn:
                conn.execute(
                    "INSERT OR IGNORE INTO categories (slug, name, url, discovered_at) VALUES (?, ?, ?, ?)",
                    (slug, name, url, datetime.now(timezone.utc).isoformat()),
                )

    def get_existing_urls(self, category: str) -> Set[str]:
        with sqlite3.connect(self.db_path, timeout=15) as conn:
            rows = conn.execute(
                "SELECT url FROM businesses WHERE category = ?", (category,)
            ).fetchall()
            return {r[0] for r in rows}

    def get_total_businesses(self) -> int:
        with sqlite3.connect(self.db_path, timeout=15) as conn:
            return conn.execute("SELECT COUNT(*) FROM businesses").fetchone()[0]

    def get_total_categories(self) -> int:
        with sqlite3.connect(self.db_path, timeout=15) as conn:
            return conn.execute("SELECT COUNT(*) FROM categories").fetchone()[0]

    def get_duplicate_stats(self) -> dict:
        """Get deduplication statistics based on fingerprints."""
        with sqlite3.connect(self.db_path, timeout=15) as conn:
            total = conn.execute("SELECT COUNT(*) FROM businesses").fetchone()[0]
            # Count unique non-empty fingerprints + records with empty fingerprint
            unique_fp = conn.execute(
                "SELECT COUNT(DISTINCT fingerprint) FROM businesses "
                "WHERE fingerprint IS NOT NULL AND fingerprint != ''"
            ).fetchone()[0]
            no_fp = conn.execute(
                "SELECT COUNT(*) FROM businesses "
                "WHERE fingerprint IS NULL OR fingerprint = ''"
            ).fetchone()[0]
            unique = unique_fp + no_fp
            return {"total": total, "unique": unique, "duplicates": total - unique}

    # ─── Checkpoints ──────────────────────────────────────────────────────

    async def checkpoint_save(self, slug: str, page: int, urls_found: int,
                              completed: bool = False):
        async with self.db_lock:
            with sqlite3.connect(self.db_path, timeout=15) as conn:
                conn.execute(
                    "INSERT OR REPLACE INTO checkpoints "
                    "(category_slug, last_page, urls_found, completed, updated_at) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (slug, page, urls_found, int(completed),
                     datetime.now(timezone.utc).isoformat()),
                )

    def checkpoint_get(self, slug: str) -> Optional[dict]:
        with sqlite3.connect(self.db_path, timeout=15) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                "SELECT * FROM checkpoints WHERE category_slug = ?", (slug,)
            ).fetchone()
            return dict(row) if row else None

    def checkpoint_summary(self) -> dict:
        with sqlite3.connect(self.db_path, timeout=15) as conn:
            completed = conn.execute(
                "SELECT COUNT(*) FROM checkpoints WHERE completed = 1 AND urls_found > 0"
            ).fetchone()[0]
            in_progress = conn.execute(
                "SELECT COUNT(*) FROM checkpoints WHERE completed = 0"
            ).fetchone()[0]
            total_cats = conn.execute(
                "SELECT COUNT(*) FROM categories"
            ).fetchone()[0]
            return {
                "completed": completed,
                "in_progress": in_progress,
                "total_categories": total_cats,
            }

    def checkpoint_clear(self):
        with sqlite3.connect(self.db_path, timeout=15) as conn:
            conn.execute("DELETE FROM checkpoints")
            conn.commit()

    # ─── Raw CSV (batched append) ─────────────────────────────────────────

    async def append_raw_csv(self, data: dict, category: str, url: str):
        """Buffer a business row and flush to CSV when buffer is full."""
        async with self.csv_lock:
            try:
                elapsed = 0.0
                if self._start_time_ref:
                    elapsed = self._start_time_ref()
                row = {
                    "Category": category,
                    "URL": url,
                    "Scraped_At": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
                    "Runtime_Sec": f"{elapsed:.0f}",
                }
                if self._uses_stable_raw_header:
                    all_keys = (
                        list(BASE_FIELDS)
                        + sorted(set(SOCIAL_DOMAINS.values()))
                        + ["First_Seen", "Last_Seen"]
                    )
                else:
                    all_keys = list(BASE_FIELDS) + sorted(
                        k for k in data if k not in BASE_FIELDS
                    )
                row.update({k: data.get(k, "") for k in all_keys})
                self._csv_buffer.append(row)

                fieldnames = ["Category", "URL", "Scraped_At", "Runtime_Sec"] + all_keys
                if self._uses_stable_raw_header:
                    self._csv_fieldnames = fieldnames
                else:
                    # Recompute fieldnames only when new keys appear
                    if self._csv_fieldnames is None or set(fieldnames) != set(self._csv_fieldnames):
                        self._csv_fieldnames = fieldnames

                if len(self._csv_buffer) >= self._csv_flush_size:
                    self._flush_csv_buffer()
            except Exception as e:
                log.warning(f"Raw CSV buffer failed: {e}")

    def _flush_csv_buffer(self):
        """Write buffered rows to CSV file (call under csv_lock)."""
        if not self._csv_buffer or not self._csv_fieldnames:
            return
        try:
            file_exists = os.path.exists(self.raw_csv_path)
            if file_exists:
                self._normalize_raw_csv(self._csv_fieldnames)
            with open(self.raw_csv_path, "a", newline="", encoding="utf-8-sig") as fh:
                writer = csv.DictWriter(fh, fieldnames=self._csv_fieldnames, extrasaction="ignore")
                if not file_exists:
                    writer.writeheader()
                writer.writerows(self._csv_buffer)
            self._csv_buffer.clear()
        except Exception as e:
            log.warning(f"Raw CSV flush failed: {e}")

    def _normalize_raw_csv(self, fieldnames: List[str]):
        try:
            with open(self.raw_csv_path, "r", newline="", encoding="utf-8-sig") as fh:
                rows = list(csv.reader(fh))
        except FileNotFoundError:
            return
        except Exception as e:
            log.warning(f"Raw CSV normalize read failed: {e}")
            return

        if not rows:
            return

        header = rows[0]
        if header == fieldnames and all(len(row) == len(fieldnames) for row in rows[1:]):
            return

        normalized = [fieldnames]
        for row in rows[1:]:
            if header == fieldnames:
                normalized.append((row + [""] * len(fieldnames))[:len(fieldnames)])
                continue

            values = {header[i]: value for i, value in enumerate(row[:len(header)]) if i < len(header)}
            padded = (row + [""] * len(fieldnames))[:len(fieldnames)]
            normalized.append([
                values.get(name, padded[i] if i < len(padded) else "")
                for i, name in enumerate(fieldnames)
            ])

        try:
            with open(self.raw_csv_path, "w", newline="", encoding="utf-8-sig") as fh:
                writer = csv.writer(fh)
                writer.writerows(normalized)
        except Exception as e:
            log.warning(f"Raw CSV normalize write failed: {e}")

    async def flush_csv(self):
        """Force-flush remaining CSV buffer (call at shutdown)."""
        async with self.csv_lock:
            self._flush_csv_buffer()

    # ─── Final Export: Priority-Sorted CSVs ───────────────────────────────

    def export_sorted_csvs(self, dedup: bool = True):
        """Export per-category CSVs with priority sorting from SQLite.

        Args:
            dedup: If True, merge cross-category duplicates (by fingerprint).
                   Keeps earliest first_seen_at winner, adds Appears_In_Categories.
        """
        with sqlite3.connect(self.db_path, timeout=15) as conn:
            conn.row_factory = sqlite3.Row

            categories = [
                row["category"]
                for row in conn.execute("SELECT DISTINCT category FROM businesses")
            ]

            if not categories:
                log.warning("No data in database to export.")
                return

            social_keys = set(SOCIAL_DOMAINS.values())

            base_order = ["Name", "Phone_1", "Phone_2", "Phone_3", "WhatsApp",
                          "Email", "Location", "Area",
                          "Governorate", "Country", "About", "Website",
                          "Rating", "Views", "Fax", "First_Seen", "Last_Seen", "URL"]

            # Build cross-category fingerprint map for dedup
            fp_category_map: Dict[str, List[str]] = {}  # fingerprint → [categories]
            fp_first_seen: Dict[str, str] = {}  # fingerprint → earliest first_seen_at
            if dedup:
                for row in conn.execute(
                    "SELECT fingerprint, category, first_seen_at FROM businesses "
                    "WHERE fingerprint IS NOT NULL AND fingerprint != '' "
                    "ORDER BY COALESCE(first_seen_at, scraped_at) ASC"
                ):
                    fp = row["fingerprint"]
                    cat = row["category"]
                    if fp not in fp_category_map:
                        fp_category_map[fp] = []
                        fp_first_seen[fp] = row["first_seen_at"] or ""
                    if cat not in fp_category_map[fp]:
                        fp_category_map[fp].append(cat)

            for cat in categories:
                rows = conn.execute(
                    "SELECT url, data, first_seen_at, last_seen_at, fingerprint "
                    "FROM businesses WHERE category = ?", (cat,)
                ).fetchall()

                records = []
                all_keys: Set[str] = set()
                for row in rows:
                    d = json.loads(row["data"])
                    d["URL"] = row["url"]
                    d["First_Seen"] = (row["first_seen_at"] or "")[:19]
                    d["Last_Seen"] = (row["last_seen_at"] or "")[:19]
                    fp = row["fingerprint"] or ""
                    if dedup and fp and fp in fp_category_map:
                        cats_list = fp_category_map[fp]
                        d["Appears_In_Categories"] = ", ".join(cats_list)
                    all_keys.update(d.keys())
                    records.append(d)

                if not records:
                    continue

                fieldnames = [c for c in base_order if c in all_keys]
                if "Appears_In_Categories" in all_keys:
                    fieldnames.append("Appears_In_Categories")
                fieldnames += sorted(k for k in all_keys if k not in fieldnames)

                # Priority sort: businesses needing digital presence ranked first
                def sort_tier(r):
                    has_web = bool(r.get("Website"))
                    has_soc = any(r.get(k) for k in social_keys)
                    if not has_web and not has_soc:
                        return 1  # Highest priority: no online presence
                    if not has_web and has_soc:
                        return 2  # Has social but no website
                    return 3      # Has website

                records.sort(key=sort_tier)

                safe_name = re.sub(r"[^a-z0-9]+", "_", cat.lower()).strip("_")
                filename = os.path.join(self.categories_dir, f"{safe_name}.csv")

                with open(filename, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    for rec in records:
                        writer.writerow({col: rec.get(col, "") for col in fieldnames})

                log.info(f"Exported {len(records)} records -> {filename} (priority sorted)")

            # Combined file (with dedup)
            all_rows = conn.execute(
                "SELECT url, category, data, first_seen_at, last_seen_at, fingerprint "
                "FROM businesses ORDER BY COALESCE(first_seen_at, scraped_at) ASC"
            ).fetchall()
            if all_rows:
                all_records = []
                all_keys_combined: Set[str] = set()
                seen_fps: Dict[str, dict] = {}  # fingerprint → winner record
                total_before = len(all_rows)

                for row in all_rows:
                    fp = row["fingerprint"] or ""
                    # Dedup: merge fields from duplicate into winner record
                    if dedup and fp:
                        if fp in seen_fps:
                            # Merge: fill empty fields in winner from this duplicate
                            winner = seen_fps[fp]
                            dup_data = json.loads(row["data"])
                            for k, v in dup_data.items():
                                if v and not winner.get(k, ""):
                                    winner[k] = v
                            continue
                        # First occurrence becomes the winner

                    d = json.loads(row["data"])
                    d["URL"] = row["url"]
                    d["Category"] = row["category"]
                    d["First_Seen"] = (row["first_seen_at"] or "")[:19]
                    d["Last_Seen"] = (row["last_seen_at"] or "")[:19]
                    if dedup and fp and fp in fp_category_map:
                        d["Appears_In_Categories"] = ", ".join(fp_category_map[fp])
                    if dedup and fp:
                        seen_fps[fp] = d  # store ref for field merging from later dupes
                    all_keys_combined.update(d.keys())
                    all_records.append(d)

                # After merging, update all_keys from enriched records
                for rec in all_records:
                    all_keys_combined.update(rec.keys())

                combined_order = ["Category"] + [c for c in base_order if c in all_keys_combined]
                if "Appears_In_Categories" in all_keys_combined:
                    combined_order.append("Appears_In_Categories")
                combined_order += sorted(k for k in all_keys_combined if k not in combined_order)

                with open(self.all_csv_path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.DictWriter(f, fieldnames=combined_order)
                    writer.writeheader()
                    for rec in all_records:
                        writer.writerow({col: rec.get(col, "") for col in combined_order})

                dedup_msg = ""
                if dedup:
                    removed = total_before - len(all_records)
                    dedup_msg = f" ({removed} cross-category duplicates merged)" if removed else ""
                log.info(
                    f"Exported {len(all_records)} unique records -> "
                    f"{self.all_csv_path}{dedup_msg}"
                )

    # ─── Validation + Raw-CSV Fallback ────────────────────────────────────

    @staticmethod
    def _csv_row_count(path: str) -> int:
        """Count data rows (excluding header) in a CSV file. Returns 0 if missing."""
        if not path or not os.path.exists(path):
            return 0
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                # Subtract header line; clamp to >= 0 for empty files
                n = sum(1 for _ in f)
                return max(0, n - 1)
        except Exception:
            return 0

    def validate_exports(self) -> dict:
        """Compute row counts across raw CSV, all CSV, per-category CSVs, and DB.

        Returns a dict: {raw, all, categories, db, db_unique}.
        `db` is the raw business row count; `db_unique` is the deduped count
        (unique fingerprints + rows with no fingerprint), which is what the
        combined `all CSV` should match when `dedup=True`.
        """
        raw_n = self._csv_row_count(self.raw_csv_path)
        all_n = self._csv_row_count(self.all_csv_path)
        cats_n = 0
        if os.path.isdir(self.categories_dir):
            for fn in os.listdir(self.categories_dir):
                if fn.lower().endswith(".csv"):
                    cats_n += self._csv_row_count(
                        os.path.join(self.categories_dir, fn)
                    )
        db_n = self.get_total_businesses()
        try:
            stats = self.get_duplicate_stats()
            db_unique = int(stats.get("unique", db_n))
        except Exception:
            db_unique = db_n
        return {
            "raw": raw_n,
            "all": all_n,
            "categories": cats_n,
            "db": db_n,
            "db_unique": db_unique,
        }

    def export_from_raw_csv(self, dedup: bool = True):
        """Rebuild per-category sorted CSVs and the combined all CSV directly
        from the raw CSV file. Used as a backup path when DB-based export
        produces inconsistent output (e.g. DB corrupted or empty).

        Dedup keeps the latest occurrence per URL (raw CSV is append-only,
        so refreshes for the same URL appear later in the file).
        """
        if not os.path.exists(self.raw_csv_path):
            log.warning("[raw-fallback] Raw CSV not found; cannot rebuild.")
            return

        try:
            with open(self.raw_csv_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        except Exception as e:
            log.error(f"[raw-fallback] Failed to read raw CSV: {e}")
            return

        if not rows:
            log.warning("[raw-fallback] Raw CSV is empty.")
            return

        # Per-category dedup: keep latest row for each (URL, Category) pair so
        # refreshed scrapes supersede earlier ones, but a URL that legitimately
        # appears in multiple categories is preserved in each.
        if dedup:
            by_pair: Dict[tuple, dict] = {}
            for r in rows:
                u = r.get("URL", "")
                c = r.get("Category", "")
                if u:
                    by_pair[(u, c)] = r
            rows = list(by_pair.values())

        # Drop transient runtime columns from output
        drop_cols = {"Runtime_Sec", "Scraped_At"}

        social_keys = set(SOCIAL_DOMAINS.values())
        base_order = ["Name", "Phone_1", "Phone_2", "Phone_3", "WhatsApp",
                      "Email", "Location", "Area",
                      "Governorate", "Country", "About", "Website",
                      "Rating", "Views", "Fax", "First_Seen", "Last_Seen", "URL"]

        def sort_tier(r):
            has_web = bool(r.get("Website"))
            has_soc = any(r.get(k) for k in social_keys)
            if not has_web and not has_soc:
                return 1
            if not has_web and has_soc:
                return 2
            return 3

        # ── Per-category CSVs ────────────────────────────────────────────
        by_cat: Dict[str, List[dict]] = {}
        for r in rows:
            cat = r.get("Category", "")
            if cat:
                by_cat.setdefault(cat, []).append(r)

        os.makedirs(self.categories_dir, exist_ok=True)

        for cat, recs in by_cat.items():
            if not recs:
                continue
            all_keys: Set[str] = set()
            for r in recs:
                all_keys.update(k for k, v in r.items() if v)
            all_keys -= drop_cols
            all_keys.discard("Category")
            fieldnames = [c for c in base_order if c in all_keys]
            fieldnames += sorted(k for k in all_keys if k not in fieldnames)

            recs_sorted = sorted(recs, key=sort_tier)

            safe_name = re.sub(r"[^a-z0-9]+", "_", cat.lower()).strip("_")
            filename = os.path.join(self.categories_dir, f"{safe_name}.csv")
            try:
                with open(filename, "w", newline="", encoding="utf-8-sig") as fh:
                    writer = csv.DictWriter(
                        fh, fieldnames=fieldnames, extrasaction="ignore"
                    )
                    writer.writeheader()
                    for rec in recs_sorted:
                        writer.writerow({col: rec.get(col, "") for col in fieldnames})
                log.info(f"[raw-fallback] {len(recs_sorted)} -> {filename}")
            except Exception as e:
                log.warning(f"[raw-fallback] Failed writing {filename}: {e}")

        # ── Combined all CSV (URL-deduped, with merged categories) ───────
        # Mirror the DB-based combined export: collapse cross-category
        # duplicates by URL, merge non-empty fields from each occurrence,
        # and aggregate the category list under `Appears_In_Categories`.
        if dedup:
            combined_by_url: Dict[str, dict] = {}
            cats_by_url: Dict[str, List[str]] = {}
            for r in rows:
                u = r.get("URL", "")
                if not u:
                    continue
                cat = r.get("Category", "")
                if u not in combined_by_url:
                    combined_by_url[u] = dict(r)
                    cats_by_url[u] = [cat] if cat else []
                else:
                    winner = combined_by_url[u]
                    for k, v in r.items():
                        if v and not winner.get(k):
                            winner[k] = v
                    if cat and cat not in cats_by_url[u]:
                        cats_by_url[u].append(cat)
            for u, rec in combined_by_url.items():
                if len(cats_by_url[u]) > 1:
                    rec["Appears_In_Categories"] = ", ".join(cats_by_url[u])
            combined_rows = list(combined_by_url.values())
        else:
            combined_rows = rows

        all_keys_combined: Set[str] = set()
        for r in combined_rows:
            all_keys_combined.update(k for k, v in r.items() if v)
        all_keys_combined -= drop_cols
        combined_order = ["Category"] + [c for c in base_order if c in all_keys_combined]
        if "Appears_In_Categories" in all_keys_combined:
            combined_order.append("Appears_In_Categories")
        combined_order += sorted(
            k for k in all_keys_combined if k not in combined_order
        )

        # Apply the same priority sort as per-category output for consistency
        combined_rows = sorted(combined_rows, key=sort_tier)

        try:
            with open(self.all_csv_path, "w", newline="", encoding="utf-8-sig") as fh:
                writer = csv.DictWriter(
                    fh, fieldnames=combined_order, extrasaction="ignore"
                )
                writer.writeheader()
                for rec in combined_rows:
                    writer.writerow({col: rec.get(col, "") for col in combined_order})
            log.info(f"[raw-fallback] {len(combined_rows)} -> {self.all_csv_path}")
        except Exception as e:
            log.warning(f"[raw-fallback] Failed writing {self.all_csv_path}: {e}")

    def safe_export(self, dedup: bool = True) -> dict:
        """Primary DB-based export with validation and automatic raw-CSV fallback.

        Flow:
          1. Run `export_sorted_csvs` (DB-based).
          2. Validate row counts (raw / all / categories / db / db_unique).
          3. If `all` or `categories` are clearly inconsistent (zero while
             data exists, or far below the deduped DB count), rebuild from
             the raw CSV.
          4. Re-validate and log the final counts.

        Serialised by `_export_lock` so concurrent callers (e.g. periodic
        + final shutdown export) cannot trample each other's CSV writes.

        Returns the final counts dict.
        """
        with self._export_lock:
            try:
                self.export_sorted_csvs(dedup=dedup)
            except Exception as e:
                log.error(f"[validate] Primary export failed: {e}")

            counts = self.validate_exports()
            log.info(
                "[validate] db=%d (unique=%d)  all=%d  categories=%d  raw=%d"
                % (counts["db"], counts["db_unique"], counts["all"],
                   counts["categories"], counts["raw"])
            )

            db_n = counts["db"]
            db_unique = counts.get("db_unique", db_n)
            all_n = counts["all"]
            cats_n = counts["categories"]
            raw_n = counts["raw"]

            # Expected size of the combined CSV when dedup=True is db_unique;
            # when dedup=False it is db_n. Compare against the right baseline.
            expected_all = db_unique if dedup else db_n

            # Inconsistency heuristics — fallback only if raw clearly has data
            # that didn't make it to the export artifacts.
            needs_fallback = False
            reason = ""
            if raw_n > 0 and all_n == 0:
                needs_fallback = True
                reason = "all CSV is empty"
            elif raw_n > 0 and cats_n == 0 and db_n > 0:
                needs_fallback = True
                reason = "no per-category CSVs written"
            elif expected_all > 0 and all_n > 0 and all_n * 2 < expected_all:
                # All CSV has < half of expected unique rows — suspicious.
                needs_fallback = True
                reason = (
                    f"all CSV ({all_n}) far below expected ({expected_all})"
                )

            if needs_fallback:
                log.warning(
                    f"[validate] Inconsistency: {reason}. Rebuilding from raw CSV…"
                )
                self.export_from_raw_csv(dedup=dedup)
                counts = self.validate_exports()
                log.info(
                    "[validate] post-fallback  all=%d  categories=%d  raw=%d"
                    % (counts["all"], counts["categories"], counts["raw"])
                )

            return counts

    # ─── Excel Export ─────────────────────────────────────────────────────

    def export_excel(self, output_path: Optional[str] = None, dedup: bool = True):
        """Export data to formatted .xlsx with summary, per-category, and all-data sheets.

        Args:
            output_path: Path for the .xlsx file. Defaults to {country}_all.xlsx.
            dedup: Merge cross-category duplicates in the combined sheet.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            log.warning("openpyxl not installed — skipping Excel export")
            return

        if output_path is None:
            output_path = self.all_csv_path.replace(".csv", ".xlsx")

        with sqlite3.connect(self.db_path, timeout=15) as conn:
            conn.row_factory = sqlite3.Row

            categories = [
                row["category"]
                for row in conn.execute("SELECT DISTINCT category FROM businesses")
            ]
            if not categories:
                log.warning("No data for Excel export.")
                return

            social_keys = set(SOCIAL_DOMAINS.values())
            base_order = ["Name", "Phone_1", "Phone_2", "Phone_3", "WhatsApp",
                          "Email", "Location", "Area", "Governorate", "Country",
                          "About", "Website", "Rating", "Views", "Fax",
                          "First_Seen", "Last_Seen", "URL"]

            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            link_font = Font(color="0563C1", underline="single", size=10)
            thin_border = Border(
                bottom=Side(style="thin", color="E0E0E0"),
            )

            # Build fingerprint map for dedup
            fp_category_map: Dict[str, List[str]] = {}
            if dedup:
                for row in conn.execute(
                    "SELECT fingerprint, category FROM businesses "
                    "WHERE fingerprint IS NOT NULL AND fingerprint != '' "
                    "ORDER BY COALESCE(first_seen_at, scraped_at) ASC"
                ):
                    fp = row["fingerprint"]
                    cat = row["category"]
                    if fp not in fp_category_map:
                        fp_category_map[fp] = []
                    if cat not in fp_category_map[fp]:
                        fp_category_map[fp].append(cat)

            wb = Workbook()

            # ── Summary sheet ─────────────────────────────────
            ws_summary = wb.active
            ws_summary.title = "Summary"
            total = conn.execute("SELECT COUNT(*) FROM businesses").fetchone()[0]
            dup_stats = self.get_duplicate_stats()
            cat_count = len(categories)

            summary_data = [
                ("Metric", "Value"),
                ("Total Businesses", str(total)),
                ("Unique Businesses", str(dup_stats["unique"])),
                ("Duplicates Removed", str(dup_stats["duplicates"])),
                ("Categories", str(cat_count)),
                ("", ""),
                ("Category", "Count"),
            ]
            for cat in categories:
                cnt = conn.execute(
                    "SELECT COUNT(*) FROM businesses WHERE category = ?", (cat,)
                ).fetchone()[0]
                summary_data.append((cat.replace("_", " ").title(), str(cnt)))

            for row_idx, (col_a, col_b) in enumerate(summary_data, 1):
                ws_summary.cell(row=row_idx, column=1, value=col_a)
                ws_summary.cell(row=row_idx, column=2, value=col_b)
                if row_idx == 1 or row_idx == 7:
                    for col in (1, 2):
                        cell = ws_summary.cell(row=row_idx, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
            ws_summary.column_dimensions["A"].width = 30
            ws_summary.column_dimensions["B"].width = 15

            # ── Per-category sheets ───────────────────────────
            for cat in categories:
                rows = conn.execute(
                    "SELECT url, data, first_seen_at, last_seen_at, fingerprint "
                    "FROM businesses WHERE category = ?", (cat,)
                ).fetchall()
                if not rows:
                    continue

                records = []
                all_keys: Set[str] = set()
                for row in rows:
                    d = json.loads(row["data"])
                    d["URL"] = row["url"]
                    d["First_Seen"] = (row["first_seen_at"] or "")[:19]
                    d["Last_Seen"] = (row["last_seen_at"] or "")[:19]
                    all_keys.update(d.keys())
                    records.append(d)

                fieldnames = [c for c in base_order if c in all_keys]
                fieldnames += sorted(k for k in all_keys if k not in base_order)

                # Priority sort
                def sort_tier(r, _social=social_keys):
                    has_web = bool(r.get("Website"))
                    has_soc = any(r.get(k) for k in _social)
                    if not has_web and not has_soc:
                        return 1
                    if not has_web and has_soc:
                        return 2
                    return 3
                records.sort(key=sort_tier)

                # Sheet name: sanitize and truncate
                sheet_name = re.sub(r"[\\/*?\[\]:]", "", cat.replace("_", " ").title())
                sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
                # Deduplicate sheet names
                existing_names = [ws.title for ws in wb.worksheets]
                if sheet_name in existing_names:
                    for i in range(2, 100):
                        candidate = f"{sheet_name[:28]}_{i}"
                        if candidate not in existing_names:
                            sheet_name = candidate
                            break

                ws = wb.create_sheet(title=sheet_name)
                self._write_excel_sheet(ws, fieldnames, records, header_font,
                                        header_fill, alt_fill, link_font, thin_border)

            # ── All Data sheet (with dedup) ───────────────────
            all_rows = conn.execute(
                "SELECT url, category, data, first_seen_at, last_seen_at, fingerprint "
                "FROM businesses ORDER BY COALESCE(first_seen_at, scraped_at) ASC"
            ).fetchall()

            if all_rows:
                all_records = []
                all_keys_combined: Set[str] = set()
                seen_fps: Set[str] = set()

                for row in all_rows:
                    fp = row["fingerprint"] or ""
                    if dedup and fp:
                        if fp in seen_fps:
                            continue
                        seen_fps.add(fp)

                    d = json.loads(row["data"])
                    d["URL"] = row["url"]
                    d["Category"] = row["category"]
                    d["First_Seen"] = (row["first_seen_at"] or "")[:19]
                    d["Last_Seen"] = (row["last_seen_at"] or "")[:19]
                    if dedup and fp and fp in fp_category_map:
                        d["Appears_In_Categories"] = ", ".join(fp_category_map[fp])
                    all_keys_combined.update(d.keys())
                    all_records.append(d)

                combined_order = ["Category"] + [c for c in base_order if c in all_keys_combined]
                if "Appears_In_Categories" in all_keys_combined:
                    combined_order.append("Appears_In_Categories")
                combined_order += sorted(k for k in all_keys_combined if k not in combined_order)

                ws_all = wb.create_sheet(title="All Data")
                self._write_excel_sheet(ws_all, combined_order, all_records, header_font,
                                        header_fill, alt_fill, link_font, thin_border)

            # Save atomically via temp file
            tmp_path = output_path + ".tmp"
            try:
                wb.save(tmp_path)
                os.replace(tmp_path, output_path)
                log.info(f"Excel exported -> {output_path}")
            except PermissionError:
                log.error(
                    f"Cannot write {output_path} — file may be open in Excel. "
                    "Close it and try again."
                )
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
            except Exception as e:
                log.error(f"Excel export failed: {e}")
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass

    @staticmethod
    def _write_excel_sheet(ws, fieldnames, records, header_font, header_fill,
                           alt_fill, link_font, thin_border):
        """Write records to an Excel worksheet with formatting."""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment

        # Header row
        for col_idx, col_name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        link_cols = {"URL", "Email", "Website"}
        for row_idx, rec in enumerate(records, 2):
            for col_idx, col_name in enumerate(fieldnames, 1):
                value = rec.get(col_name, "")
                # Sanitize formula injection
                if isinstance(value, str) and value and value[0] in "=+-@":
                    value = "'" + value

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Make URLs/emails clickable
                if col_name in link_cols and value and not value.startswith("'"):
                    if col_name == "Email" and "@" in value:
                        cell.hyperlink = f"mailto:{value}"
                        cell.font = link_font
                    elif col_name in ("URL", "Website") and value.startswith("http"):
                        cell.hyperlink = value
                        cell.font = link_font

            # Alternate row shading
            if row_idx % 2 == 0:
                for col_idx in range(1, len(fieldnames) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = alt_fill

        # Auto-width columns
        for col_idx, col_name in enumerate(fieldnames, 1):
            max_len = len(col_name)
            for row_idx in range(2, min(len(records) + 2, 102)):  # Sample first 100 rows
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

        # Freeze header + auto-filter
        ws.freeze_panes = "A2"
        if records:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(records) + 1}"
